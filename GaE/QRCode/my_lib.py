from tkinter import *
import shutil
import pyqrcode
from fpdf import FPDF
from tkinter import filedialog
from tkinter import messagebox
import openpyxl
from openpyxl.styles import PatternFill

# Check Free SN => Which cell sheet is not in RED color?
def checkFreeSn(worksheet, listbox):
    listbox.delete(0, 'end')
    Gp = []
    Rev = []
    Sn = []
    Mac1 = []
    Mac2 = []
    list_cnt = 0 # to count the number of QR codes added to the listbox
    row_init = 3; # skiip the three first lines because they are titles.
    for i in range(row_init, worksheet.max_row):
        cell_color = worksheet["A"+str(i)].fill.start_color.rgb # get cell color
        #print(cell_color)
        if (cell_color == "00000000" and worksheet["A"+str(i)].value is None):
            #print(worksheet["A"+str(i)].value)
            break
        elif cell_color == "00000000":
            #Gp.append(str(worksheet["A" + str(i)].value))
            Gp.append(str(worksheet["A" + str(i)].value))
            Rev.append(str(worksheet["B" + str(i)].value))
            Sn.append(str(worksheet["C" + str(i)].value))
            if str(worksheet) == '<Worksheet "CCU_SoC_Ctrl & MAC">':
                Mac1.append(str(worksheet["D" + str(i)].value))
                Mac2.append(str(worksheet["E" + str(i)].value))
                # insert item to the CCU_SoC_Ctrl listbox
                listbox.insert(END, "QR" + str(list_cnt+1) + "=> " + Gp[list_cnt] + "; Rev.: " + Rev[list_cnt] +
                              "; Sn: " + Sn[list_cnt] + "; MAC1: " + Mac1[list_cnt] + "; MAC2: " +
                               Mac2[list_cnt])
                #generate temp png images for qr codes
                qrparam1 = 'Gamesa Electric\n\r'
                qrparam2 = 'CCU_SoC_Ctrl\n\r'
                qrparam3 = "GP: " + Gp[list_cnt] + "\n\r"  # ctrlG
                qrparam4 = "Rev.: " + Rev[list_cnt] + "\n\r"  # ctrlRev
                qrparam5 = "SN: " + Sn[list_cnt] + "\n\r"  # ctrlSn
                qrparam6 = "MAC ETH1: " + Mac1[list_cnt] + "\n\r"  # ctrlMac1
                qrparam7 = "MAC ETH2: " + Mac2[list_cnt] + "\n\r"  # ctrlMac2
                qr_code = pyqrcode.create(qrparam1 + qrparam2 + qrparam3 + qrparam4 + qrparam5 + qrparam6 + qrparam7)
                qr_code.png('./temp/qrcode_ctrl{}.png'.format(list_cnt), scale=2)
            else:
                listbox.insert(END, "QR" + str(list_cnt+1) + "=> " + Gp[list_cnt] + "; Rev.: " + Rev[list_cnt] +
                               "; Sn: " + Sn[list_cnt])
                # generate temp png images for qr codes
                if str(worksheet) == '<Worksheet "CCU_S4d_Adapt">':
                    qrparam1 = 'Gamesa Electric\n\r'
                    qrparam2 = 'CCU_S4d_Adapt\n\r'
                    qrparam3 = "GP: " + Gp[list_cnt] + "\n\r"  # s4dGp
                    qrparam4 = "Rev.: " + Rev[list_cnt] + "\n\r"  # s4dRev
                    qrparam5 = "SN: " + Sn[list_cnt] + "\n\r"  # s4dSn
                    qr_code = pyqrcode.create(qrparam1 + qrparam2 + qrparam3 + qrparam4 + qrparam5)
                    qr_code.png('./temp/qrcode_s4d{}.png'.format(list_cnt), scale=2)
                else:
                    qrparam1 = 'Gamesa Electric\n\r'
                    qrparam2 = 'CCU_SoC\n\r'
                    qrparam3 = "GP: " + Gp[list_cnt] + "\n\r"  # s4dGp
                    qrparam4 = "Rev.: " + Rev[list_cnt] + "\n\r"  # s4dRev
                    qrparam5 = "SN: " + Sn[list_cnt] + "\n\r"  # s4dSn
                    qr_code = pyqrcode.create(qrparam1 + qrparam2 + qrparam3 + qrparam4 + qrparam5)
                    qr_code.png('./temp/qrcode_soc{}.png'.format(list_cnt), scale=2)
            list_cnt += 1

    return list_cnt

#commandImport: Import excel file to fill the entries
def commandImport(ctrlList, s4dList, socList):

    # Global var used to store the excel file in a temp memory
    global fileExcel
    # Global vars to count the number of Sn printed
    global SnCntFree_SoC
    global SnCntFree_Ctrl
    global SnCntFree_S4d

    # Open excel file
    fileExcel = filedialog.askopenfilename(title="Abrir", initialdir="./", filetypes=(("Ficheros de Excel", "*.xlsx"),
                                        ("Todos los ficheros", "*.*")))

    if not fileExcel: # handle cancel button filedialog.askopenfilename
        return

    # Open Excel File
    wb = openpyxl.load_workbook(fileExcel, data_only=True) # data_only=True to read the cell data instead of the formula
    # Open Sheets
    wsSoC = wb['CCU_SoC']
    wsS4d = wb['CCU_S4d_Adapt']
    wsCtrl = wb['CCU_SoC_Ctrl & MAC']
    # Check free SN + fill listbox
    SnCntFree_Ctrl = checkFreeSn(wsCtrl, ctrlList)
    SnCntFree_S4d = checkFreeSn(wsS4d, s4dList)
    SnCntFree_SoC = checkFreeSn(wsSoC, socList)

    # check the number of SN read.
    if SnCntFree_SoC > 5:
        messagebox.showwarning("Warning Info", "The number of SN in CCU_SoC excel tab > 5."
                                               "lease edit the Excel file to assure SN <= 5")
    if SnCntFree_Ctrl > 5:
        messagebox.showwarning("Warning Info", "The number of SN in CCU_SoC_Ctrl excel tab > 5."
                                               "lease edit the Excel file to assure SN <= 5")
    if SnCntFree_S4d > 5:
        messagebox.showwarning("Warning Info", "The number of SN in CCU_S4d_Adapt excel tab > 5."
                                               "lease edit the Excel file to assure SN <= 5")

    if SnCntFree_SoC != SnCntFree_Ctrl != SnCntFree_S4d:
        messagebox.showwarning("Warning Info", "There are different number of SN in the excel tabs."
                                               " Every tab must have the same number of SN."
                                                " Please edit the excel file to assure it.")


#commandPrint: Print Ctrl QR codes on the screen
def commandQrgenCtrl(event, labelQr):
    widget = event.widget
    selection = widget.curselection()
    # Refresh QR image
    image_new = PhotoImage(file='./temp/qrcode_ctrl{}.png'.format(selection[0])) #selection[0] convert tuple to int
    labelQr.configure(image=image_new, bd=5, relief="groove")
    labelQr.photo = image_new

#commandPrint: Print S4d QR codes on the screen
def commandQrgenS4d(event, labelQr):
    widget = event.widget
    selection = widget.curselection()
    # Refresh QR image
    image_new = PhotoImage(file='./temp/qrcode_s4d{}.png'.format(selection[0]))
    labelQr.configure(image=image_new, bd=5, relief="groove")
    labelQr.photo = image_new

#commandPrint: Print S4d QR codes on the screen
def commandQrgenSoc(event, labelQr):
    widget = event.widget
    selection = widget.curselection()
    # Refresh QR image
    image_new = PhotoImage(file='./temp/qrcode_soc{}.png'.format(selection[0]))
    labelQr.configure(image=image_new, bd=5, relief="groove")
    labelQr.photo = image_new

#commandExport: Export to PDF QR codes
def commandExport():
    pdf = FPDF()
    pdf.add_page()

    # Set font type and line width for pdf titles and drawings
    pdf.set_font("Arial", size=12)
    pdf.set_line_width(1)
    #  printing vertical lines
    pdf.line(x1=70, y1=5, x2=70, y2=290)
    pdf.line(x1=135, y1=5, x2=135, y2=290)

    ygap = 0 # gap between QR code lines
    for i in range(SnCntFree_SoC): # maximum 5 qr codes can be printed in one page
        # priting horizontal lines
        if i < 4: # dont print last horizontal line
            pdf.line(x1=5, y1=50+ygap, x2=200, y2=50+ygap)
        # printing page headers
        pdf.text(x=20, y=5+ygap, txt="CCU_SoC_ctrl (QR{})".format(i+1))
        pdf.text(x=85, y=5+ygap, txt="CCU_s4d_adapt (QR{})".format(i+1))
        pdf.text(x=155, y=5+ygap, txt="CCU_SoC (QR{})".format(i+1))
        #CCU_SoC_ctrl
        pdf.image('./temp/qrcode_ctrl{}.png'.format(i), x=25,  y=20+ygap, w=10, h=10)
        pdf.image('./temp/qrcode_ctrl{}.png'.format(i), x=45, y=20+ygap, w=10, h=10)
        # CCU_s4d_adapt
        pdf.image('./temp/qrcode_s4d{}.png'.format(i), x=100,  y=20+ygap, w=10, h=10)
        # CCU_SoC
        pdf.image('./temp/qrcode_soc{}.png'.format(i), x=160,  y=20+ygap, w=10, h=10)
        ygap += 60

    filePdf = filedialog.asksaveasfile(mode="w", defaultextension=".pdf", filetypes=[("Ficheros PDF", "*.pdf")])

    if filePdf is None: # asksaveasfile return `None` if dialog closed with "cancel".
        return

    pdf.output(filePdf.name, "F")
    pdf.close

    fillRedXls()

    messagebox.showinfo("GaE windows messagebox", "Fichero con codigos QR generado con exito")

# Define an action whhen the program is closed
def on_closing(root):
    shutil.rmtree("./temp")
    root.destroy()

def fillRedXls():
    # Open Excel File
    wb = openpyxl.load_workbook(fileExcel, data_only=True)  # data_only=True to read the cell data instead of the formula
    # Open Sheets
    wsSoC = wb['CCU_SoC']
    wsS4d = wb['CCU_S4d_Adapt']
    wsCtrl = wb['CCU_SoC_Ctrl & MAC']
    row_init=3 # the three first rows of the excel file are tittles.
    for i in range(row_init, wsSoC.max_row):
        cell_color = wsSoC["A"+str(i)].fill.start_color.rgb # get cell color
        if (cell_color == "00000000" and wsSoC["A"+str(i)].value is None): #check if the cell is empty
            #print(worksheet["A"+str(i)].value)
            break
        elif cell_color == "00000000": #check the background color of the cell (check cell white color)
            redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
            # Red fill CCU_SoC excel tab
            wsSoC["A" + str(i)].fill = redFill
            wsSoC["B" + str(i)].fill = redFill
            wsSoC["C" + str(i)].fill = redFill
            # Red fill CCU_S4d_Adapt excel tab
            wsS4d["A" + str(i)].fill = redFill
            wsS4d["B" + str(i)].fill = redFill
            wsS4d["C" + str(i)].fill = redFill
            # Red fill CCU_S4d_Adapt excel tab
            wsCtrl["A" + str(i)].fill = redFill
            wsCtrl["B" + str(i)].fill = redFill
            wsCtrl["C" + str(i)].fill = redFill
            wsCtrl["D" + str(i)].fill = redFill
            wsCtrl["E" + str(i)].fill = redFill

    wb.save('MAC&CCU_SN_Mod.xlsx')


'''
    row_init=3
    for i in range(row_init, worksheet.max_row):
        cell_color = worksheet["A"+str(i)].fill.start_color.rgb # get cell color
        #print(cell_color)
        if (cell_color == "00000000" and worksheet["A"+str(i)].value is None):
            #print(worksheet["A"+str(i)].value)
            break
        elif cell_color == "00000000":


    wb = openpyxl.load_workbook('MAC&CCU_SN.xlsx')
    #ws = wb.active
    ws = wb["CCU_SoC"]
    redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    i = 8
    ws["A"+str(i)].fill = redFill
    wb.save('test.xlsx')    
'''
