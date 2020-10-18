import openpyxl
from openpyxl.styles import Color, PatternFill


excelFileName = "./database.xlsx"
wb = openpyxl.load_workbook(excelFileName, data_only=True)
ws = wb["DB"]
#print(ws["K5"].value)
# fill the cell with a color
redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')
whiteFill = PatternFill(start_color='FFFFFFFF',
                   end_color='FFFFFFFF',
                   fill_type='solid')                   
ws["K5"].fill = redFill
#ws["K5"].fill = whiteFill
wb.save("database.xlsx")
# get the color cell
cell_color = ws['K5'].fill.start_color.index
print(cell_color)

