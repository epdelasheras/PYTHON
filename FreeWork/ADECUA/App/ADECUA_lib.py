from tkinter import *
import cv2
from PIL import Image, ImageTk
import openpyxl

# root window attributes
WIN_SIZE = "1366x768"
WIN_TITLE = "ADECUA"
WIN_ICO_PATH = ".\pics\ADECUA.ico"

# Pic sizes
WIDTH_FLATPIC = 700
HEIGHT_FLATPIC = 500
WIDTH_FLOORPIC = 380
HEIGHT_FLOORPIC = 380

# Tree items
NAMEBLOCK = "Portal "
NAMEFLOOR = "Planta "
NAMEROOM = " Habitacion/es"
NAMEVAR = "Variante "

# Floor pic areas
BLK1_X1 = 40
BLK1_X2 = 125
BLK1_Y1 = 115
BLK1_Y2 = 135

# to identify the area selected on the Floor pic.
def areaId(label_floor_x, label_floor_y):
	if (label_floor_x > BLK1_X1 and label_floor_x < BLK1_X2 and label_floor_y > BLK1_Y1 and label_floor_y < BLK1_Y2):
		return AREA_ID[0]

# make a rectangle on the floor image
def highlightArea(label_floor_x, label_floor_y):
	floor_picCv2 = cv2.imread("./pics/floorviews/PlantaBaja.png") # load image in open cv format
	floor_picCv2Copy = floor_picCv2.copy() # create a copy of the previous image
    # draw a rectangle over the image
	x, y, w, h = 10, 10, 10, 10  # Rectangle parameters
	cv2.rectangle(floor_picCv2Copy, (x, y), (x + w, y + h), (0, 200, 0), -1)  # A filled rectangle
	alpha = 0.4  # Transparency factor.
	floor_picCv2Mod = cv2.addWeighted(floor_picCv2Copy, alpha, floor_picCv2, 1 - alpha, 0)  # image with the rectangle

    # OpenCV represents images in BGR order; however PIL represents
    # images in RGB order, so we need to swap the channels
	b, g, r = cv2.split(floor_picCv2Mod)
	floor_picCv2_RGB = cv2.merge((r, g, b))

	# Convert the Image object into a TkPhoto object
	floor_pic = Image.fromarray(floor_picCv2_RGB)

	floor_picCopy = floor_pic.copy()
	floor_picCopyResize = floor_picCopy.resize((WIDTH_FLOORPIC, HEIGHT_FLOORPIC), Image.ANTIALIAS)
	floor_picTk = ImageTk.PhotoImage(image=floor_picCopyResize)
	return floor_picTk

def addItemsTreeview(Treeview):
    # adding elements to the tree. The previous tree label is indicated
    # to have an event response

	wb = openpyxl.load_workbook("./database.xlsx", data_only=True)  # data_only=True to read the cell data instead of the formula
	# read all ws and create a list
	lst_blk = []
	for sheet in wb:
		lst_blk.append(sheet.title)

	# Add blocks to the treeview.
	for i in range(len(lst_blk)):
		Treeview.insert("", END, text=NAMEBLOCK + str(i + 1), iid=lst_blk[i], tags=("mytag",))

	# Add
	for sheet in range(len(lst_blk)):
		ws = wb[lst_blk[sheet]]
		# Adding floors to the treeview
		lst_floor = []
		for i in range(2, ws.max_row+1):
			lst_floor.append(str(ws["A"+str(i)].value))
		lst_floor_mod = list(set(lst_floor))  # remove duplicate items
		lst_floor_mod.sort()  # sort list items
		for i in range(len(lst_floor_mod)):
			Treeview.insert(lst_blk[sheet], END, text=NAMEFLOOR + str(i),
							iid=lst_blk[sheet] + lst_floor_mod[i], tags=("mytag",))


		# Adding rooms to the treeview
		lst_cnt_floor = []
		for i in range(len(lst_floor_mod)):
			cnt_floor = 0
			for j in range(len(lst_floor)):
				if lst_floor_mod[i] == lst_floor[j]:
					cnt_floor += 1
			lst_cnt_floor.append(cnt_floor)

		cnt_room = 0 # counter to follow the type of rooms in the excel file.
		for i in range(len(lst_cnt_floor)):
			lst_room = []
			for j in range(lst_cnt_floor[i]):
				lst_room.append(str(ws["B" + str(j+2+cnt_room)].value))
			cnt_room = cnt_room + j + 1
			lst_room_mod = list(set(lst_room))  # remove duplicate items
			lst_room_mod.sort()  # sort list items
			for k in range(len(lst_room_mod)):
				n_room = lst_room_mod[k] #Get the number of room from the excel.
				#print(n_room[4:5])
				Treeview.insert(lst_blk[sheet] + lst_floor_mod[i], END, text=n_room[4:5] + NAMEROOM,
								iid=lst_blk[sheet] + lst_floor_mod[i] + lst_room_mod[k], tags=("mytag",))

		# Adding variants to the treeview
		lst_var = []
		lst_flat = []
		for i in range(2, ws.max_row + 1):
			lst_var.append(str(ws["C" + str(i)].value))
			lst_flat.append(str(ws["B" + str(i)].value))
		for i in range(len(lst_var)):
			n_var = lst_var[i]
			Treeview.insert(lst_blk[sheet] + lst_floor[i] + lst_flat[i], END, text=NAMEVAR + n_var[3:4],
								iid=lst_blk[sheet] + lst_floor[i] + lst_flat[i] + lst_var[i], tags=("mytag",))