from tkinter import *
from tkinter import ttk
from PIL import Image, ImageTk
import openpyxl

# root window attributes
WIN_SIZE = "1400x800"
WIN_TITLE = "ADECUA"
PIC_EXTENSION = ".jpg"
PIC_PATH = "./pics/"

# Pic sizes
WIDTH_FLATPIC = 700
HEIGHT_FLATPIC = 500
WIDTH_FLOORPIC = 380
HEIGHT_FLOORPIC = 380

# Tree items
N_ITEMS = 4 # when tipology tree view is selected, tree_view_split = 4 items
WS_ROW_START = 5
NAMESTRUCTURE = "Estructura "
NAMEPROFILE = "Perfil "
NAMEFLOOR = "Planta "

def searchLocation(worksheet, tree_item):
    # It is used in "loadPic" method to search the coordinates of the flat.

    wb = openpyxl.load_workbook("./database.xlsx", data_only=True)  # data_only=True to read the cell data instead of the formula
    ws = wb[worksheet]
    for i in range(WS_ROW_START, ws.max_row + 1):
        if (ws["K" + str(i)].value == tree_item):
            return ws["H" + str(i)].value

def addItemsTreeview(Treeview):
    # adding elements to the tree. The previous tree label is indicated
    # to have an event response

    wb = openpyxl.load_workbook("./database.xlsx", data_only=True)  # data_only=True to read the cell data instead of the formula
    # read all ws and create a list
    lst_ws = []
    for sheet in wb:
        lst_ws.append(sheet.title)

    # tuples to return at the end of the method
    ws_fname = []
    ws_fplace = []
    ws_ftipo = []

    # Loop to read all worsheets of the excel book
    for n_sheet in range(len(lst_ws)):

        # saving a copy in memory of the current excel worksheet to process
        ws = wb[lst_ws[n_sheet]]

        # read all ws rows and create a list with all the ws file names.
        lst_fname = []
        lst_fplace = []
        lst_ftipo = []
        for i in range(WS_ROW_START, ws.max_row + 1):
            if (ws["K" + str(i)].value != None):  # skip empty rows
                lst_fname.append(str(ws["K" + str(i)].value)) # list with file names
                lst_fplace.append(str(ws["H" + str(i)].value)) # list with coordinates
                lst_ftipo.append(str(ws["F" + str(i)].value)) # list with tiplogies

        # Split every file name string to do an analysis later
        lst_fname_split = []
        for i in range(len(lst_fname)):
            lst_fname_split.append(lst_fname[i].split("-"))

        # create a list with every element of every lst_fname_split item
        lst_struct_prof = []
        lst_struct_prof_floor = []
        for i in range(len(lst_fname_split)):
            str_split = lst_fname_split[i]
            lst_struct_prof.append(str_split[0] + "-" + str_split[1])
            lst_struct_prof_floor.append(str_split[0] + "-" + str_split[1]
                                     + "-" + str_split[2])

        # Adding structures to the treeview
        Treeview.insert("", END, text=NAMESTRUCTURE + lst_ws[n_sheet],
                        iid=lst_ws[n_sheet], tags=("mytag",))

        # Adding profiles to the treeview
        print(lst_struct_prof)
        lst_struct_prof_sort = list(set(lst_struct_prof))
        lst_struct_prof_sort.sort()
        for i in range(len(lst_struct_prof_sort)):
            profile = lst_struct_prof_sort[i]
            profile_split = profile.split("-")
            print(profile)
            print(profile_split)
            Treeview.insert(profile_split[0], END, text=NAMEPROFILE + profile_split[1],
                    iid=profile, tags=("mytag",))

        # Adding floors to the treeview
        lst_struct_prof_floor_sort = list(set(lst_struct_prof_floor))
        lst_struct_prof_floor_sort.sort()
        for i in range(len(lst_struct_prof_floor_sort)):
            floor = lst_struct_prof_floor_sort[i]
            floor_split = floor.split("-")
            Treeview.insert(floor_split[0] + "-" + floor_split[1], END,
                        text=NAMEFLOOR + floor_split[2], iid=floor,
                        tags=("mytag",))

        # Adding types to the treeview
        for i in range(len(lst_fname)):
            type = lst_fname[i]
            type_split = type.split("-")
            Treeview.insert(type_split[0] + "-" + type_split[1] + "-" + type_split[2], END, text=type_split[3], iid=type, tags=("mytag",))

        # Create a list with all the filenames to be used in other parts of the program
        ws_fname.append(lst_fname)

        # Create a list with all the coordinates to be used in other parts of the program
        ws_fplace.append(lst_fplace)

        # Create a list with all the tipologies to be used in other parts of the program
        ws_ftipo.append(lst_ftipo)

    return ws_fname, ws_fplace, ws_ftipo


def addItemsListboxRoom(Listbox, file_names):
    # Adding items to the listbox

    # Figure out which is the max. number of rooms in the database
    n_room = []
    for i in range(len(file_names)): #num max. of rows
        for j in range(len(file_names[i])): #num max of cols
            if len(re.findall(r"[0-9]+D", file_names[i][j])) > 0: #skip flat with no rooms
                tipology_room = re.findall(r"[0-9]+D", file_names[i][j])
                if len(tipology_room) > 1:
                    # check for this tipology construction: 3MF_0D.A_DA+3MF_2D.A1_DA
                    room_add = str(int(tipology_room[0][:1]) + int(tipology_room[1][:1])) + "D"
                    n_room.append(room_add)
                else:
                    # check for this tipology construction: 3MF_1D.B
                    n_room.append(tipology_room[0])

    # removing duplicated items of the listbox
    n_room = list(set(n_room))
    n_room.sort()

    # Adding rooms to the listbox
    for i in range(len(n_room)):
        Listbox.insert(i, str(i+1) + " Dormitorio/s")

    return n_room