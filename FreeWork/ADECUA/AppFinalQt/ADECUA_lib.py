import re
from typing import List

import openpyxl
from openpyxl.styles import Color, PatternFill
from PyQt5 import QtCore, QtGui, QtWidgets, Qt
from tinydb import TinyDB, Query

# root window attributes
MAINWIN_WIDTH = 1400
MAINWIN_HEIGHT = 800
WIN_TITLE = "ADECUA"
PIC_EXTENSION = ".jpg"
PIC_PATH = "./pics/"

# Pic sizes
WIDTH_FLATPIC = 842
HEIGHT_FLATPIC = 595
WIDTH_FLOORPIC = 380
HEIGHT_FLOORPIC = 380

# Tree items
N_ITEMS = 4 # when tipology tree widget is selected, tree_view_split = 4 items
WS_ROW_START = 5
NAMESTRUCTURE = "Estructura "
NAMEPROFILE = "Perfil "
NAMEFLOOR = "Planta "


def AdecuaDefGuiConfig(tb_room, flat_pic, tree_widget, lb_room, excelFileName):
# method to load the default configuration in ADECUA main window
    # Customization of the label which shows Nroom and coodinates
    tb_room.setStyleSheet(" font-size: 14px; qproperty-alignment: AlignCenter; "
                               "border: 1px solid black; ")
    tb_room.setText("Coordenadas|NºHabitaciones")
    # load default image in flat_pic label
    load_pic = QtGui.QPixmap(PIC_PATH + "Diseño Base y Variante 1_2.jpg")
    load_pic = load_pic.scaled(WIDTH_FLATPIC, HEIGHT_FLATPIC,
                               #QtCore.Qt.IgnoreAspectRatio,
                               QtCore.Qt.KeepAspectRatio,
                               QtCore.Qt.SmoothTransformation)
    flat_pic.setPixmap(load_pic)
    # Add number of rooms to choose to the listbox and connect lb_room with mouse click
    lboxRoomAddItems(lb_room, excelFileName)
    # Add tree items and connect tree_widget with mouse click
    qtwidget_struct, qtwidget_profile, qtwidget_floor, qtwidget_type, tree_struct, tree_profile, \
    tree_floor, tree_type, tree_picname = treewidgetAddItems(tree_widget, excelFileName)   
    # return variables
    return qtwidget_struct, qtwidget_profile, qtwidget_floor, qtwidget_type, tree_struct, \
           tree_profile, tree_floor, tree_type, tree_picname 

def WindowPopUpWarning(text):
# shoe dialog pop up warning window 
    dialog = QtWidgets.QMessageBox()
    dialog.setWindowTitle(WIN_TITLE)
    dialog.setWindowIcon(QtGui.QIcon(PIC_PATH+WIN_TITLE+PIC_EXTENSION))
    dialog.setIcon(QtWidgets.QMessageBox.Warning)
    dialog.setText(text)
    dialog.addButton(QtWidgets.QMessageBox.Ok)
    dialog.exec() 

def lboxRoomAddItems(Listbox, excelFileName):
# Adding items to the listbox
    wb = openpyxl.load_workbook(excelFileName, data_only=True)  # data_only=True to read the cell data instead of the formula
    # read all ws and create a list
    lst_ws = []
    for sheet in wb:
        lst_ws.append(sheet.title)
    # create list with number of rooms
    lst_nrooms = []
    # create a var to figure out if there are LCXX types in the excel file
    lc_cnt = int(0)
    # Loop to read all worsheets of the excel book
    for n_sheet in range(len(lst_ws)):
        # saving a copy in memory of the current excel worksheet to process
        ws = wb[lst_ws[n_sheet]]
        for i in range(WS_ROW_START, ws.max_row + 1):
            if re.match(r"[0-9]", str(ws["I" + str(i)].value)) != None:
                lst_nrooms.append(ws["I" + str(i)].value)
            elif str(ws["I" + str(i)].value) == "-" :
                lc_cnt += 1
    # set and sort list
    lst_nrooms_sort = list(set(lst_nrooms))
    lst_nrooms_sort.sort()
    # Adding rooms to the listbox
    for i in range(len(lst_nrooms_sort)):
        Listbox.insertItem(i, str(i+1) + " Dormitorio/s")
    # check for the LCxx types
    if lc_cnt > 0:
        Listbox.insertItem(i+1, "- Dormitorio/s")

def lboxRoomPlaceAddItems(item, Listbox, excelFileName):
# Add tipologies and coordinates to lb_roomplace
    wb = openpyxl.load_workbook(excelFileName, data_only=True)  # data_only=True to read the cell data instead of the formula
    # read all ws and create a list
    lst_ws = []
    for sheet in wb:
        lst_ws.append(sheet.title)
    Listbox.clear()  # delete listbox items before add new ones.
    # print(item[:1])
    # list to save items read from the excel file
    lst_item = []
    # Loop to read all worsheets of the excel book
    for n_sheet in range(len(lst_ws)):
        # saving a copy in memory of the current excel worksheet to process
        ws = wb[lst_ws[n_sheet]]
        for i in range(WS_ROW_START, ws.max_row + 1):
            #print(str(ws["I" + str(i)].value))
            if item[:1] == str(ws["I" + str(i)].value):
                lst_item.append(ws["K" + str(i)].value)
    # add sorting items to the listbox
    lst_item.sort()
    for i in range(len(lst_item)):
        Listbox.insertItem(i, lst_item[i])

def excelLockPicname(picname, excelFileName):
# lock picname cell in excel file
    # block flat filling with red color the picname cell 
    picname_split = picname.split("-", 1)
    #print(picname_split)
    #print(picname_split[0])                
    wb = openpyxl.load_workbook(excelFileName, data_only=True)
    ws = wb[picname_split[0]]
    redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')
    for i in range(WS_ROW_START, ws.max_row + 1):
        if (ws["K" + str(i)].value == picname):
            ws["K" + str(i)].fill = redFill
            #print(ws["K" + str(i)].value)    
    wb.save(excelFileName)

def excelUnlockPicname(picname, excelFileName):
# Unlock picname cell in excel file
    # Unblock flat filling with original cell color the picname (white) 
    picname_split = picname.split("-", 1)
    #print(picname_split)
    #print(picname_split[0])                
    wb = openpyxl.load_workbook(excelFileName, data_only=True)
    ws = wb[picname_split[0]]
    whiteFill = PatternFill(start_color='FFFFFFFF',
                   end_color='FFFFFFFF',
                   fill_type='solid')          
    for i in range(WS_ROW_START, ws.max_row + 1):
        if (ws["K" + str(i)].value == picname):
            ws["K" + str(i)].fill = whiteFill
            #print(ws["K" + str(i)].value)    
    wb.save(excelFileName)               
        
def treewidgetAddItems(Treewidget, excelFileName):
# adding elements to the tree. The previous tree label is indicated
# to have an event response
    wb = openpyxl.load_workbook(excelFileName, data_only=True)  # data_only=True to read the cell data instead of the formula
    # read all ws and create a list
    lst_ws = []
    for sheet in wb:
        lst_ws.append(sheet.title)
    #Create main Treewidget column
    Treewidget.setColumnCount(1)
    Treewidget.setHeaderLabel("ARBOL DE SELECCIÓN")
    # Create tuplas with all treewidgets
    lst_struct_widget = []
    lst_profile_widget = []
    lst_floor_widget = []
    lst_type_widget = []
    # Create tuplas with the text of all treewidgets
    lst_struct_txt = []
    lst_profile_txt = []
    lst_floor_txt = []
    lst_type_txt = []
    lst_picname = []
    # Loop to read all worsheets of the excel book
    for n_sheet in range(len(lst_ws)):
        # saving a copy in memory of the current excel worksheet to process
        ws = wb[lst_ws[n_sheet]]
        # read all ws rows and create a list with all the ws file names.
        lst_fname = []
        lst_fplace = []
        lst_ftipo = []
        for i in range(WS_ROW_START, ws.max_row + 1):
            if (ws["K" + str(i)].value != None):  # skip empty rows
                lst_fname.append(str(ws["K" + str(i)].value)) # list with file names
                lst_fplace.append(str(ws["H" + str(i)].value)) # list with coordinates
                lst_ftipo.append(str(ws["F" + str(i)].value)) # list with tiplogies
        # Split every file name string to do an analysis later
        lst_fname_split = []
        for i in range(len(lst_fname)):
            lst_fname_split.append(lst_fname[i].split("-"))
        # create a list with every element of every lst_fname_split item
        lst_struct_prof = []
        lst_struct_prof_floor = []
        for i in range(len(lst_fname_split)):
            str_split = lst_fname_split[i]
            lst_struct_prof.append(str_split[0] + "-" + str_split[1])
            lst_struct_prof_floor.append(str_split[0] + "-" + str_split[1]
                                     + "-" + str_split[2])
        # Adding structures to the Treewidget
        tree_struct = Qt.QTreeWidgetItem(Treewidget, [NAMESTRUCTURE + lst_ws[n_sheet]])
        lst_struct_widget.append(tree_struct) # add item widget to tupla
        lst_struct_txt.append(NAMESTRUCTURE + lst_ws[n_sheet])  # add widget text to tupla
        # Adding profiles to the structures
        lst_struct_prof_sort = list(set(lst_struct_prof))
        lst_struct_prof_sort.sort()
        lst_struct_prof_floor_sort = list(set(lst_struct_prof_floor))
        lst_struct_prof_floor_sort.sort()
        for i in range(len(lst_struct_prof_sort)):
            profile = lst_struct_prof_sort[i]
            profile_split = profile.split("-")
            tree_profile = Qt.QTreeWidgetItem(tree_struct, [NAMEPROFILE + profile_split[1]])
            lst_profile_widget.append(tree_profile)  # add item widget to tupla
            lst_profile_txt.append(NAMEPROFILE + profile_split[1]) # add widget text to tupla
            for j in range(len(lst_struct_prof_floor_sort)):
                floor = lst_struct_prof_floor_sort[j]
                floor_split = floor.split("-")
                if re.fullmatch(profile_split[1], floor_split[1]) != None:
                    tree_floors = Qt.QTreeWidgetItem(tree_profile, [NAMEFLOOR + floor_split[2]])
                    lst_floor_widget.append(tree_floors)  # add item widget to tupla
                    lst_floor_txt.append(NAMEFLOOR + floor_split[2])  # add widget text to tupla
                    for k in range(len(lst_struct_prof_floor)):
                        type = lst_fname[k]
                        type_split = type.split("-")
                        if re.fullmatch(floor, lst_struct_prof_floor[k]) != None:
                            #print(type_split)
                            tree_type = Qt.QTreeWidgetItem(tree_floors, [type_split[3]])
                            lst_type_widget.append(tree_type) # add item widget to tupla
                            lst_type_txt.append(type_split[3])  # add widget text to tupla
                            lst_picname.append(type)   
    # return variables
    return lst_struct_widget, lst_profile_widget, lst_floor_widget, lst_type_widget, \
           lst_struct_txt, lst_profile_txt, lst_floor_txt, lst_type_txt, lst_picname

def treeWidgetExpandItem(tree_struct, tree_profile, tree_floor, tree_type, qtwidget_struct, qtwidget_profile,
                   qtwidget_floor,qtwidget_type, tree_widget, item_sel):
#It is used to expand a specific item on the tree_widget.
    item_sel_split = item_sel.split("-")
    #print(item_sel_split)
    item_struct = NAMESTRUCTURE + item_sel_split[0]
    item_profile = NAMEPROFILE + item_sel_split[1]
    item_floor = NAMEFLOOR + item_sel_split[2]
    item_type = item_sel_split[3]
    #print(item_struct)
    #print(item_profile)
    #print(item_floor)
    # collapse the complete tree before expanding a new item
    tree_widget.collapseAll()
    # un-select all items
    for i in range(len(tree_type)):
        qtwidget_type[i].setSelected(False)
    # expanding every item on tree one by one
    for i in range(len(tree_struct)):
        if tree_struct[i] == item_struct:
            qtwidget_struct[i].setExpanded(True)
    for i in range(len(tree_profile)):
        if tree_profile[i] == item_profile:
            qtwidget_profile[i].setExpanded(True)
    for i in range(len(tree_floor)):
        if tree_floor[i] == item_floor:
            qtwidget_floor[i].setExpanded(True)
    for i in range(len(tree_type)):
        if tree_type[i] == item_type:            
            qtwidget_type[i].setSelected(True)

def treeWidgetLoadImageAndLocation(item_sel, flat_pic, tb_room, qtwidget_type, tree_picname, excelFileName):
    # It is used to load images and coordinates into flat pic label and tb_room from tree_widget widget
    for i in range(len(qtwidget_type)):
        if qtwidget_type[i] == item_sel[0]:
            picname = tree_picname[i]
            load_pic = QtGui.QPixmap(PIC_PATH + picname + ".jpg")
            load_pic = load_pic.scaled(WIDTH_FLATPIC, HEIGHT_FLATPIC,
                                       #QtCore.Qt.IgnoreAspectRatio,
                                       QtCore.Qt.KeepAspectRatio,                                       
                                       QtCore.Qt.SmoothTransformation)
            flat_pic.setPixmap(load_pic)
            picname_split = picname.split("-", 1)
            #print(picname_split)
            location, n_room = TreeWidgetGivemeNroomLocation(picname_split[0], picname, excelFileName)
            #print(location, n_room)
            tb_room.setText(str(location) + " | " + str(n_room) + " dormitorio/s")                  


def treeWidgetRightMenuActions(statusbar, tree_widget, excelFileName, db_ADECUA_TableFlatFav, 
                               db_ADECUA_TableFlatBook, db_ADECUA_TableFlatBuy, flat_picname):
# actions to execute when right click is made on tree widget item of ADECUA main window.
    # read statusbar string
    statusBarText = statusbar.currentMessage()
    # filtering the string to get the doc_id of the data base                
    statusBarText_split = statusBarText.split('[')           
    # execute code only if a client is selected.
    if len(statusBarText_split) > 1:                                  
        # save item selected from tree widget
        item_sel = tree_widget.selectedItems()        
        # save db_id
        db_id = str(statusBarText_split[1][:-1])
        # Getting info from the selected item
        #print(flat_picname)
        flat_picname_split = flat_picname.split("-", 1)
        location, n_room = TreeWidgetGivemeNroomLocation(flat_picname_split[0], flat_picname, excelFileName)
        #print(location)
        #print(n_room)                   
        if tree_widget.action == tree_widget.favAction:                        
            # improve this action doing a click filtering. just in case 
            # the use makes click several times over the same item                        
            db_ADECUA_TableFlatFav.insert({'Id': db_id, 'Picname': flat_picname,
                                            'NumRoom':n_room, 'Coordinates': location})
            #print ("Favorito")
            #print(self.db_ADECUA_TableFlatFav.all())                        
        elif tree_widget.action == tree_widget.bookAction:
            # improve this action doing a click filtering. just in case 
            # the use makes click several times over the same item    
            db_ADECUA_TableFlatBook.insert({'Id': db_id, 'Picname': flat_picname,
                                            'NumRoom':n_room, 'Coordinates': location})
            #print ("Reserva")
            # disable this item and make it not selectable
            for i in item_sel:
                i.setDisabled(True)
                i.setSelected(False)
            # lock item in the excel file
            excelLockPicname(flat_picname, excelFileName)   
        elif tree_widget.action == tree_widget.buyAction:
            # improve this action doing a click filtering. just in case 
            # the use makes click several times over the same item    
            db_ADECUA_TableFlatBuy.insert({'Id': db_id, 'Picname': flat_picname,
                                            'NumRoom':n_room, 'Coordinates': location})
            #print ("Compra")
            # disable this item and make it not selectable
            for i in item_sel:
                i.setDisabled(True)
                i.setSelected(False)
            # lock item in the excel file
            excelLockPicname(flat_picname, excelFileName)

def TreeWidgetGivemeNroomLocation(worksheet, tree_item, excelFileName):
# It is used in "loadPic" method to search the coordinates of the flat.
    wb = openpyxl.load_workbook(excelFileName, data_only=True)  # data_only=True to read the cell data instead of the formula
    ws = wb[worksheet]
    for i in range(WS_ROW_START, ws.max_row + 1):
        if (ws["K" + str(i)].value == tree_item):
            return ws["H" + str(i)].value, ws["I" + str(i)].value

def treeWidgetItemLock(tree_struct, tree_profile, tree_floor, tree_type, qtwidget_struct, qtwidget_profile,
                      qtwidget_floor,qtwidget_type, tree_widget, item_sel):
#It is used to lock a specific item on the tree_widget.        
    item_sel_split = item_sel.split("-")
    #print(item_sel_split)
    item_struct = NAMESTRUCTURE + item_sel_split[0]
    item_profile = NAMEPROFILE + item_sel_split[1]
    item_floor = NAMEFLOOR + item_sel_split[2]
    item_type = item_sel_split[3]
    #print(item_struct)
    #print(item_profile)
    #print(item_floor)
    # collapse the complete tree before expanding a new item
    tree_widget.collapseAll()
    # un-select all items
    for i in range(len(tree_type)):
        qtwidget_type[i].setSelected(False)
    # expanding every item on tree one by one
    for i in range(len(tree_struct)):
        if tree_struct[i] == item_struct:
            qtwidget_struct[i].setExpanded(True)
    for i in range(len(tree_profile)):
        if tree_profile[i] == item_profile:
            qtwidget_profile[i].setExpanded(True)
    for i in range(len(tree_floor)):
        if tree_floor[i] == item_floor:
            qtwidget_floor[i].setExpanded(True)
    for i in range(len(tree_type)):
        if tree_type[i] == item_type:            
            qtwidget_type[i].setDisabled(True)
            qtwidget_type[i].setSelected(False)
            #print("item locked")    
                   
def treeWidgetItemUnLock(tree_struct, tree_profile, tree_floor, tree_type, qtwidget_struct, 
                         qtwidget_profile, qtwidget_floor,qtwidget_type, tree_widget, item_sel):
#It is used to unlock a specific item on the tree_widget.
    item_sel_split = item_sel.split("-")
    #print(item_sel_split)
    item_struct = NAMESTRUCTURE + item_sel_split[0]
    item_profile = NAMEPROFILE + item_sel_split[1]
    item_floor = NAMEFLOOR + item_sel_split[2]
    item_type = item_sel_split[3]
    #print(item_struct)
    #print(item_profile)
    #print(item_floor)
    # collapse the complete tree before expanding a new item
    tree_widget.collapseAll()
    # un-select all items
    for i in range(len(tree_type)):
        qtwidget_type[i].setSelected(False)
    # expanding every item on tree one by one
    for i in range(len(tree_struct)):
        if tree_struct[i] == item_struct:
            qtwidget_struct[i].setExpanded(True)
    for i in range(len(tree_profile)):
        if tree_profile[i] == item_profile:
            qtwidget_profile[i].setExpanded(True)
    for i in range(len(tree_floor)):
        if tree_floor[i] == item_floor:
            qtwidget_floor[i].setExpanded(True)
    for i in range(len(tree_type)):
        if tree_type[i] == item_type:            
            qtwidget_type[i].setSelected(True)
            qtwidget_type[i].setDisabled(False)
            #print("iem unlocked")